"""Known-number IDOC lookup from the Processed Apps List spreadsheet.

Provides:
- Loading the ground truth spreadsheet into a name→number / number→name index
- Minimum digit length validation (5+ digits)
- Fuzzy matching OCR candidates against the known set
- Name-based fallback lookup when OCR fails entirely

All IDOC numbers in the 2026 spreadsheet are 5–6 digits (range ~11000–166000).
"""
from __future__ import annotations

import logging
import re
from pathlib import Path
from typing import Any

from rising_sun.identity import normalize_person_name, person_name_key

logger = logging.getLogger(__name__)

_DEFAULT_XLSX = (
    Path(__file__).resolve().parent.parent.parent
    / "IDOC"
    / "Data"
    / "1. Processed Apps List.xlsx"
)

MIN_IDOC_DIGITS = 5
MAX_IDOC_DIGITS = 6


class IdocDirectory:
    """In-memory index of known IDOC numbers ↔ names from the spreadsheet."""

    def __init__(self, xlsx_path: Path | str | None = None) -> None:
        self._by_number: dict[str, str] = {}          # "154753" → "Armstrong, Adria Lillian"
        self._by_name_key: dict[tuple[str, str], list[str]] = {}  # (first, last) → [numbers]
        self._all_numbers: set[str] = set()

        path = Path(xlsx_path) if xlsx_path else _DEFAULT_XLSX
        if path.exists():
            self._load(path)
        else:
            logger.warning("IDOC spreadsheet not found at %s", path)

    @property
    def known_numbers(self) -> set[str]:
        return self._all_numbers

    def _load(self, path: Path) -> None:
        import openpyxl
        import re

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            if sheet_name.lower() in {"recommendations"}:
                continue
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=2, values_only=True):
                name_raw = row[0]
                num_raw = row[1]
                if name_raw is None or num_raw is None:
                    continue

                # Handle mixed formats like "RSO/123456", "123456/RSO",
                # "Denied/100666", plain numbers, and float representations.
                num_str = str(num_raw).strip()
                number = None
                try:
                    number = str(int(num_raw))
                except (ValueError, TypeError):
                    # Try extracting digits from mixed format (e.g. "RSO/123456")
                    digits = re.findall(r"\d{5,6}", num_str)
                    if digits:
                        number = digits[0]

                if number is None:
                    continue

                name_str = str(name_raw).strip()
                self._by_number[number] = name_str
                self._all_numbers.add(number)

                key = person_name_key(name_str)
                if key:
                    self._by_name_key.setdefault(key, []).append(number)

        wb.close()
        logger.info(
            "Loaded %d known IDOC numbers from %s",
            len(self._all_numbers),
            path.name,
        )

    def lookup_by_number(self, number: str) -> str | None:
        """Return the spreadsheet name for a known number, or None."""
        return self._by_number.get(number)

    def lookup_by_name(self, name: str) -> list[str]:
        """Return known IDOC numbers matching a person name (first+last key)."""
        key = person_name_key(name)
        if key is None:
            return []
        return list(self._by_name_key.get(key, []))

    def is_known(self, number: str) -> bool:
        return number in self._all_numbers

    # ------------------------------------------------------------------
    # Fuzzy matching
    # ------------------------------------------------------------------

    def fuzzy_match(self, candidate: str) -> list[str]:
        """Generate fuzzy digit variants and return those that exist in the known set.

        Handles common OCR errors:
        - Single-digit substitution (e.g. 177632 → 127632)
        - Missing digit / insertion (e.g. 16463 → 164463)
        - Extra digit / deletion (e.g. 140464 → 1404)
        - Adjacent transpositions
        """
        matches: set[str] = set()

        # 1. Direct match
        if candidate in self._all_numbers:
            matches.add(candidate)

        # 2. Single-digit substitution
        for i in range(len(candidate)):
            for d in "0123456789":
                if d == candidate[i]:
                    continue
                variant = candidate[:i] + d + candidate[i + 1:]
                if variant in self._all_numbers:
                    matches.add(variant)

        # 3. Insert a digit at each position (catches truncation)
        for i in range(len(candidate) + 1):
            for d in "0123456789":
                variant = candidate[:i] + d + candidate[i:]
                if len(variant) <= MAX_IDOC_DIGITS and variant in self._all_numbers:
                    matches.add(variant)

        # 4. Delete a digit at each position (catches spurious insertion)
        if len(candidate) > MIN_IDOC_DIGITS:
            for i in range(len(candidate)):
                variant = candidate[:i] + candidate[i + 1:]
                if len(variant) >= MIN_IDOC_DIGITS and variant in self._all_numbers:
                    matches.add(variant)

        # 5. Adjacent transpositions
        for i in range(len(candidate) - 1):
            variant = candidate[:i] + candidate[i + 1] + candidate[i] + candidate[i + 2:]
            if variant in self._all_numbers:
                matches.add(variant)

        return sorted(matches)

    def best_match(
        self,
        candidates: list[str],
        filename_name: str,
    ) -> tuple[str | None, str]:
        """Find the best known-number match from OCR candidates.

        Tries direct + fuzzy matching, then scores by name similarity.

        Returns (best_number, method) where method describes how it was found.
        If no match, returns (None, "no_match").
        """
        # Phase 1: direct match in known set, with name verification
        name_key = person_name_key(filename_name) if filename_name else None
        direct_hit = None
        for c in candidates:
            if c in self._all_numbers:
                # Verify name matches before trusting direct OCR hit
                if name_key:
                    spreadsheet_name = self._by_number.get(c, "")
                    c_key = person_name_key(spreadsheet_name)
                    if c_key and c_key == name_key:
                        return c, "known_direct"
                    # Name mismatch — remember first hit but don't return yet
                    if direct_hit is None:
                        direct_hit = c
                else:
                    # No filename name to verify against — trust OCR
                    return c, "known_direct"

        # Phase 2: fuzzy match each candidate
        all_fuzzy: list[tuple[str, str]] = []  # (number, source_candidate)

        for c in candidates:
            for m in self.fuzzy_match(c):
                all_fuzzy.append((m, c))

        if not all_fuzzy:
            return None, "no_match"

        # Score by name match
        if name_key:
            for number, _ in all_fuzzy:
                spreadsheet_name = self._by_number.get(number, "")
                key = person_name_key(spreadsheet_name)
                if key and key == name_key:
                    return number, "known_fuzzy_name"

        # No name-confirmed fuzzy match — do NOT return an unverified guess.
        # Caller should fall through to name_fallback instead.
        return None, "no_match"

    def name_fallback(self, filename_name: str) -> tuple[str | None, str | None]:
        """Look up IDOC number purely by filename name.

        Returns (number, spreadsheet_name) or (None, None).
        """
        numbers = self.lookup_by_name(filename_name)
        if len(numbers) == 1:
            return numbers[0], self._by_number.get(numbers[0])
        if len(numbers) > 1:
            logger.warning(
                "Multiple IDOC numbers for '%s': %s — returning first",
                filename_name,
                numbers,
            )
            return numbers[0], self._by_number.get(numbers[0])
        return None, None


def is_valid_idoc_length(number: str) -> bool:
    """Check if a candidate number has a plausible IDOC digit length (5-6)."""
    return number.isdigit() and MIN_IDOC_DIGITS <= len(number) <= MAX_IDOC_DIGITS


def filter_candidates_by_length(candidates: list[str]) -> list[str]:
    """Keep only candidates that match the known 5-6 digit IDOC number shape."""
    return [candidate for candidate in candidates if is_valid_idoc_length(candidate)]
