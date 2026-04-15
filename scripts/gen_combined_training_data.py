#!/usr/bin/env python3
"""Generate combined IDOC number + applicant name crops for multi-task TrOCR training.

Produces crops from both the IDOC# field and the applicant name field,
tagged with a task prefix so the model learns both tasks.

Output format (JSONL):
    {"image": "images/...", "text": "123456", "task": "number"}
    {"image": "images/...", "text": "John Smith", "task": "name"}

Usage:
    python scripts/gen_combined_training_data.py \
        --output-dir output/combined_training_v2
"""
from __future__ import annotations

import argparse
import hashlib
import json
import logging
import sys
from pathlib import Path

import cv2
import numpy as np
import pymupdf

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "src"))

from rising_sun.identity import clean_pdf_stem_name
from rising_sun.idoc_lookup import IdocDirectory
from rising_sun.pdf import render_pdf_page

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
logger = logging.getLogger(__name__)

DATA_DIR = Path(__file__).resolve().parent.parent / "IDOC" / "Data" / "PROCESSED APPS 2026"
WORKBOOK = Path(__file__).resolve().parent.parent / "IDOC" / "Data" / "1. Processed Apps List.xlsx"

# ---------- IDOC number region crop boxes ----------
NUM_CROP_DEFAULT = (0.62, 0.24, 0.95, 0.34)
NUM_CROP_TIGHT = (0.72, 0.238, 0.90, 0.298)
NUM_CROP_MID = (0.67, 0.238, 0.92, 0.31)

# ---------- Applicant name region crop boxes ----------
NAME_CROP_TIGHT = (0.14, 0.248, 0.60, 0.282)
NAME_CROP_CONTEXT = (0.10, 0.238, 0.64, 0.290)
NAME_CROP_WIDE = (0.08, 0.235, 0.68, 0.295)


def _crop_region(img: np.ndarray, box: tuple[float, float, float, float]) -> np.ndarray:
    h, w = img.shape[:2]
    x1, y1, x2, y2 = int(box[0] * w), int(box[1] * h), int(box[2] * w), int(box[3] * h)
    return img[y1:y2, x1:x2]


def _apply_clahe(img: np.ndarray) -> np.ndarray:
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    clahe = cv2.createCLAHE(clipLimit=3.0, tileGridSize=(8, 8))
    enhanced = clahe.apply(gray)
    return cv2.cvtColor(enhanced, cv2.COLOR_GRAY2BGR)


def _apply_binary(img: np.ndarray) -> np.ndarray:
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    _, binary = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    return cv2.cvtColor(binary, cv2.COLOR_GRAY2BGR)


def _split_bucket(name: str) -> str:
    """Deterministic train/val/test split by name hash (80/10/10)."""
    h = int(hashlib.sha256(name.encode()).hexdigest()[:4], 16) % 100
    if h < 80:
        return "train"
    elif h < 90:
        return "val"
    else:
        return "test"


def _load_name_ground_truth(workbook_path: Path) -> dict[str, str]:
    """Load name ground truth from the spreadsheet. Returns {clean_name: display_name}."""
    import openpyxl

    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = wb["2026"]

    # Find the name column
    header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    name_col = None
    for i, h in enumerate(header):
        if h and "name" in str(h).lower():
            name_col = i
            break

    if name_col is None:
        logger.warning("Could not find name column in workbook")
        return {}

    names: dict[str, str] = {}
    for row in sheet.iter_rows(min_row=2):
        val = row[name_col].value
        if val and isinstance(val, str) and val.strip():
            display = val.strip()
            # Use "First Last" format (strip middle names for label)
            parts = display.split()
            if len(parts) >= 2:
                label = f"{parts[0]} {parts[-1]}"
            else:
                label = display
            # Key is the clean stem name for matching
            clean = display.lower().strip()
            names[clean] = label

    wb.close()
    return names


# Variant definitions for IDOC number crops
NUM_VARIANTS: list[tuple[str, int, tuple, str | None]] = [
    ("num_tight_225", 225, NUM_CROP_TIGHT, None),
    ("num_default_225", 225, NUM_CROP_DEFAULT, None),
    ("num_mid_225", 225, NUM_CROP_MID, None),
    ("num_tight_225_clahe", 225, NUM_CROP_TIGHT, "clahe"),
    ("num_tight_225_binary", 225, NUM_CROP_TIGHT, "binary"),
    ("num_tight_300", 300, NUM_CROP_TIGHT, None),
    ("num_default_300", 300, NUM_CROP_DEFAULT, None),
    ("num_tight_300_clahe", 300, NUM_CROP_TIGHT, "clahe"),
    ("num_tight_300_binary", 300, NUM_CROP_TIGHT, "binary"),
    ("num_tight_400", 400, NUM_CROP_TIGHT, None),
    ("num_tight_400_binary", 400, NUM_CROP_TIGHT, "binary"),
]

# Variant definitions for name crops
NAME_VARIANTS: list[tuple[str, int, tuple, str | None]] = [
    ("name_tight_225", 225, NAME_CROP_TIGHT, None),
    ("name_context_225", 225, NAME_CROP_CONTEXT, None),
    ("name_wide_225", 225, NAME_CROP_WIDE, None),
    ("name_tight_225_clahe", 225, NAME_CROP_TIGHT, "clahe"),
    ("name_tight_225_binary", 225, NAME_CROP_TIGHT, "binary"),
    ("name_tight_300", 300, NAME_CROP_TIGHT, None),
    ("name_context_300", 300, NAME_CROP_CONTEXT, None),
    ("name_tight_300_clahe", 300, NAME_CROP_TIGHT, "clahe"),
    ("name_tight_300_binary", 300, NAME_CROP_TIGHT, "binary"),
    ("name_tight_400", 400, NAME_CROP_TIGHT, None),
    ("name_tight_400_binary", 400, NAME_CROP_TIGHT, "binary"),
]


def generate_crops_for_pdf(
    pdf_path: Path,
    idoc_number: str | None,
    name_label: str | None,
    images_dir: Path,
) -> list[dict]:
    """Generate crop variants for IDOC number and/or name and return manifest rows."""
    stem = pdf_path.stem.replace(" ", "_")
    rows = []
    page_cache: dict[int, np.ndarray] = {}

    # IDOC number crops
    if idoc_number:
        for variant_name, dpi, crop_box, augmentation in NUM_VARIANTS:
            if dpi not in page_cache:
                page_cache[dpi] = render_pdf_page(pdf_path, dpi=dpi, page_number=0)
            page_img = page_cache[dpi]
            crop = _crop_region(page_img, crop_box)
            if augmentation == "clahe":
                crop = _apply_clahe(crop)
            elif augmentation == "binary":
                crop = _apply_binary(crop)
            filename = f"{stem}__{variant_name}.png"
            cv2.imwrite(str(images_dir / filename), crop)
            rows.append({
                "image": f"images/{filename}",
                "text": idoc_number,
                "task": "number",
                "source_pdf": pdf_path.name,
                "variant": variant_name,
            })

    # Name crops
    if name_label:
        for variant_name, dpi, crop_box, augmentation in NAME_VARIANTS:
            if dpi not in page_cache:
                page_cache[dpi] = render_pdf_page(pdf_path, dpi=dpi, page_number=0)
            page_img = page_cache[dpi]
            crop = _crop_region(page_img, crop_box)
            if augmentation == "clahe":
                crop = _apply_clahe(crop)
            elif augmentation == "binary":
                crop = _apply_binary(crop)
            filename = f"{stem}__{variant_name}.png"
            cv2.imwrite(str(images_dir / filename), crop)
            rows.append({
                "image": f"images/{filename}",
                "text": name_label,
                "task": "name",
                "source_pdf": pdf_path.name,
                "variant": variant_name,
            })

    return rows


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate combined IDOC number + name training crops")
    parser.add_argument("--data-dir", type=Path, default=DATA_DIR)
    parser.add_argument("--workbook", type=Path, default=WORKBOOK)
    parser.add_argument("--output-dir", type=Path, required=True)
    parser.add_argument("--limit", type=int, default=None)
    parser.add_argument("--include-extra", nargs="*", default=[], help="Extra PDF filenames (stems) to include regardless of page count")
    args = parser.parse_args()

    output_dir: Path = args.output_dir
    images_dir = output_dir / "images"
    images_dir.mkdir(parents=True, exist_ok=True)

    # Load ground truth sources
    directory = IdocDirectory()
    name_truth = _load_name_ground_truth(args.workbook)
    logger.info(f"Loaded {len(name_truth)} name ground truth entries")

    pdfs = sorted(args.data_dir.glob("*.pdf"))
    logger.info(f"Found {len(pdfs)} total PDFs")

    # Filter to IDOC-layout PDFs (4-page, or explicitly IDOC-classified non-4-page)
    # Non-4-page PDFs are included only if listed in --include-extra
    extra_stems: set[str] = set()
    if hasattr(args, "include_extra") and args.include_extra:
        for p in args.include_extra:
            extra_stems.add(Path(p).stem)
    eligible: list[tuple[Path, str | None, str | None]] = []
    for pdf_path in pdfs:
        doc = pymupdf.open(pdf_path)
        n_pages = len(doc)
        doc.close()
        if n_pages != 4 and pdf_path.stem not in extra_stems:
            continue

        # Try to get IDOC number
        name = clean_pdf_stem_name(pdf_path)
        nums = directory.lookup_by_name(name)
        idoc_number = nums[0] if nums and len(nums[0]) >= 5 else None

        # Try to get name label
        name_label = None
        name_lower = name.lower().strip()
        if name_lower in name_truth:
            name_label = name_truth[name_lower]
        else:
            # Try matching on first+last
            parts = name.split()
            if len(parts) >= 2:
                key_fl = f"{parts[0]} {parts[-1]}".lower()
                if key_fl in name_truth:
                    name_label = name_truth[key_fl]

        # If we still don't have a name label, use the cleaned filename name
        # as "First Last" format (this is staff-typed, high quality)
        if name_label is None and len(name.split()) >= 2:
            parts = name.split()
            name_label = f"{parts[0]} {parts[-1]}"

        if idoc_number or name_label:
            eligible.append((pdf_path, idoc_number, name_label))

    logger.info(
        f"Eligible: {len(eligible)} PDFs "
        f"({sum(1 for _, n, _ in eligible if n)} with IDOC#, "
        f"{sum(1 for _, _, n in eligible if n)} with name)"
    )

    if args.limit:
        eligible = eligible[: args.limit]

    splits: dict[str, list[dict]] = {"train": [], "val": [], "test": []}

    for i, (pdf_path, idoc_number, name_label) in enumerate(eligible):
        if (i + 1) % 50 == 0 or i == 0:
            logger.info(f"[{i + 1}/{len(eligible)}] {pdf_path.name}")

        try:
            rows = generate_crops_for_pdf(pdf_path, idoc_number, name_label, images_dir)
        except Exception as e:
            logger.warning(f"Failed {pdf_path.name}: {e}")
            continue

        name = clean_pdf_stem_name(pdf_path)
        split = _split_bucket(name)
        for row in rows:
            row["split"] = split
            splits[split].append(row)

    # Write JSONL split files
    for split_name, split_rows in splits.items():
        path = output_dir / f"{split_name}.jsonl"
        with open(path, "w") as f:
            for row in split_rows:
                f.write(json.dumps({"image": row["image"], "text": row["text"], "task": row["task"]}) + "\n")
        n_num = sum(1 for r in split_rows if r["task"] == "number")
        n_name = sum(1 for r in split_rows if r["task"] == "name")
        logger.info(f"{split_name}.jsonl: {len(split_rows)} samples ({n_num} number, {n_name} name)")

    # Write full manifest
    all_rows = splits["train"] + splits["val"] + splits["test"]
    manifest_path = output_dir / "manifest.jsonl"
    with open(manifest_path, "w") as f:
        for row in all_rows:
            f.write(json.dumps(row) + "\n")

    logger.info(
        f"Done: {len(all_rows)} total crops from {len(eligible)} PDFs"
    )


if __name__ == "__main__":
    main()
